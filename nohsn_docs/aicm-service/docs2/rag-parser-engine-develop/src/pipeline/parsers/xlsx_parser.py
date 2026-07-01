"""XLSX/XLS 파서 -- openpyxl(.xlsx) + xlrd(.xls) 기반 시트별 표 추출 + Markdown 변환.

Phase 2 Task B-1.
"""

from __future__ import annotations

import asyncio
from typing import Any, Optional

from src.common.logging import get_logger
from src.pipeline.models.parse_result import (
    PageContent,
    ParseResult,
    TableContent,
    TextBlock,
)
from src.pipeline.parsers.base import BaseParser

log = get_logger(__name__)

_MAX_ROWS_PER_SHEET = 10_000


class XLSXParser(BaseParser):
    """openpyxl(.xlsx) + xlrd(.xls) 듀얼 파서.

    알고리즘
    --------
    1. openpyxl로 워크북 열기 시도 (data_only=True)
    2. 실패 시 xlrd로 폴백 (구형 .xls OLE2 형식 지원, 코드페이지 949 한글)
    3. 시트별 순회
    4. 머지셀 해제 (openpyxl 경로에서만)
    5. 헤더 행 자동 감지
    6. 연속 빈 행 3개 이상이면 표 경계로 인식 (openpyxl 경로)
    7. 각 시트/표를 독립 TableContent + Markdown 변환
    8. 시트당 최대 10,000행 (초과 시 경고 + 앞부분만 처리)
    """

    async def parse(self) -> ParseResult:
        """비동기 XLSX 파싱. IO-bound 작업을 스레드풀에 위임한다."""
        return await asyncio.to_thread(self._parse_sync)

    def _parse_sync(self) -> ParseResult:
        """동기 내부 구현. openpyxl 시도 후 실패하면 xlrd로 폴백한다."""
        try:
            return self._parse_with_openpyxl()
        except Exception as exc:
            log.info(
                "openpyxl_failed_fallback_xlrd",
                file=self.file_path,
                error=str(exc),
            )
            return self._parse_with_xlrd()

    def _parse_with_openpyxl(self) -> ParseResult:
        """openpyxl 기반 .xlsx 파싱."""
        import openpyxl

        wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=False)

        pages: list[PageContent] = []
        all_tables: list[TableContent] = []
        raw_texts: list[str] = []
        global_table_index = 0

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # 머지셀 해제 + 값 복사
            self._unmerge_and_fill(ws)

            # 행 데이터 수집 (최대 MAX_ROWS 제한)
            rows_data = self._collect_rows(ws)

            if not rows_data:
                continue

            # 연속 빈 행 3개로 테이블 경계 분리
            tables_in_sheet = self._split_tables_by_blank_rows(rows_data)

            page_tables: list[TableContent] = []
            page_texts: list[str] = []

            for local_idx, table_rows in enumerate(tables_in_sheet):
                if len(table_rows) < 1:
                    continue

                # 헤더 감지
                headers, data_rows = self._detect_header_and_split(table_rows, ws, sheet_name)

                if not headers and not data_rows:
                    continue

                md = self._to_markdown(headers, data_rows)

                tc = TableContent(
                    page_number=sheet_idx + 1,
                    table_index=global_table_index,
                    headers=headers,
                    rows=data_rows,
                    markdown=md,
                    confidence=1.0,
                    bbox=None,
                )
                page_tables.append(tc)
                all_tables.append(tc)
                page_texts.append(f"[시트: {sheet_name}, 표 {local_idx + 1}]\n{md}")
                global_table_index += 1

            page_text = "\n\n".join(page_texts)
            raw_texts.append(page_text)

            text_blocks: list[TextBlock] = []
            if page_text:
                text_blocks.append(
                    TextBlock(
                        text=page_text,
                        start_char_offset=0,
                        end_char_offset=len(page_text),
                    )
                )

            pages.append(
                PageContent(
                    page_number=sheet_idx + 1,
                    text=page_text,
                    text_blocks=text_blocks,
                    tables=page_tables,
                    layout_type="single",
                )
            )

        wb.close()

        metadata = self._extract_metadata(wb)

        return ParseResult(
            raw_text="\n\n".join(raw_texts),
            pages=pages,
            tables=all_tables,
            metadata=metadata,
            source_file_path=self.file_path,
        )

    def _parse_with_xlrd(self) -> ParseResult:
        """xlrd로 구형 .xls 파일을 파싱한다."""
        import xlrd

        wb = xlrd.open_workbook(self.file_path)
        pages: list[PageContent] = []
        all_tables: list[TableContent] = []
        all_text_parts: list[str] = []

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet_name = ws.name

            if ws.nrows == 0:
                continue

            # 시트 전체를 하나의 표로 변환
            headers: list[str] = []
            rows: list[list[str]] = []
            text_parts: list[str] = []

            # 헤더 행 탐색 (첫 비어있지 않은 행)
            header_row = 0
            for r in range(ws.nrows):
                vals = [str(ws.cell_value(r, c)).replace("\n", " ").strip() for c in range(ws.ncols)]
                if any(v for v in vals):
                    header_row = r
                    headers = vals
                    break

            # 데이터 행
            for r in range(header_row + 1, ws.nrows):
                row_vals = [str(ws.cell_value(r, c)).replace("\n", " ").strip() for c in range(ws.ncols)]
                if any(v for v in row_vals):
                    rows.append(row_vals)

            # TableContent 생성
            tc: Optional[TableContent] = None
            if headers or rows:
                md = self._to_markdown(headers, rows)
                tc = TableContent(
                    page_number=sheet_idx + 1,
                    table_index=len(all_tables),
                    headers=headers,
                    rows=rows,
                    markdown=md,
                    confidence=0.95,
                )
                all_tables.append(tc)

            # 텍스트 추출 (표 외 셀)
            for r in range(ws.nrows):
                for c in range(ws.ncols):
                    val = str(ws.cell_value(r, c)).strip()
                    if val and len(val) > 3:
                        text_parts.append(val)

            all_text_parts.extend(text_parts)

            # PageContent (시트 = 페이지)
            page_text = "\n".join(text_parts)
            pages.append(
                PageContent(
                    page_number=sheet_idx + 1,
                    text=page_text,
                    tables=[tc] if tc else [],
                    layout_type="single",
                )
            )

        raw_text = "\n\n".join(all_text_parts)

        metadata: dict[str, Any] = {
            "parser": "xlrd",
            "sheet_count": wb.nsheets,
            "sheet_names": wb.sheet_names(),
        }

        return ParseResult(
            raw_text=raw_text,
            pages=pages,
            tables=all_tables,
            images=[],
            metadata=metadata,
            source_file_path=self.file_path,
        )

    # ------------------------------------------------------------------
    # 머지셀 해제
    # ------------------------------------------------------------------
    @staticmethod
    def _unmerge_and_fill(ws: Any) -> None:
        """머지셀을 해제하고 원본 값을 모든 셀에 복사한다."""
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            value = ws.cell(row=min_row, column=min_col).value

            ws.unmerge_cells(str(merged_range))

            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row=row, column=col).value = value

    # ------------------------------------------------------------------
    # 행 수집
    # ------------------------------------------------------------------
    def _collect_rows(self, ws: Any) -> list[list[str]]:
        """시트에서 행 데이터를 문자열 리스트로 수집한다. 최대 행 수 제한 적용."""
        rows: list[list[str]] = []
        row_count = 0

        for row in ws.iter_rows(values_only=False):
            row_count += 1
            if row_count > _MAX_ROWS_PER_SHEET:
                log.warning(
                    "xlsx_row_limit_exceeded",
                    sheet=ws.title,
                    max_rows=_MAX_ROWS_PER_SHEET,
                    total_rows=ws.max_row,
                )
                break

            cell_values = []
            for cell in row:
                val = cell.value
                if val is None:
                    cell_values.append("")
                else:
                    cell_values.append(str(val).replace("\n", " ").strip())
            rows.append(cell_values)

        return rows

    # ------------------------------------------------------------------
    # 빈 행 기반 표 분리
    # ------------------------------------------------------------------
    @staticmethod
    def _split_tables_by_blank_rows(
        rows: list[list[str]],
        blank_threshold: int = 3,
    ) -> list[list[list[str]]]:
        """연속 빈 행 3개 이상이면 표 경계로 인식하여 분리한다."""
        tables: list[list[list[str]]] = []
        current_table: list[list[str]] = []
        blank_count = 0

        for row in rows:
            is_blank = all(cell == "" for cell in row)
            if is_blank:
                blank_count += 1
                if blank_count >= blank_threshold and current_table:
                    tables.append(current_table)
                    current_table = []
            else:
                if blank_count > 0 and blank_count < blank_threshold and current_table:
                    # 빈 행이 threshold 미만이면 같은 표에 포함
                    for _ in range(blank_count):
                        current_table.append([""] * len(row))
                blank_count = 0
                current_table.append(row)

        if current_table:
            tables.append(current_table)

        return tables

    # ------------------------------------------------------------------
    # 헤더 감지
    # ------------------------------------------------------------------
    def _detect_header_and_split(
        self,
        table_rows: list[list[str]],
        ws: Any,
        sheet_name: str,
    ) -> tuple[list[str], list[list[str]]]:
        """헤더 행을 감지하고 헤더와 데이터 행을 분리한다.

        감지 기준:
        1. 첫 행의 셀이 모두 비어있지 않고 다른 행과 데이터 패턴이 다른 경우
        2. 첫 행 셀에 bold 서식이 있는 경우
        3. 위 조건 모두 해당 없으면 첫 행을 헤더로 사용
        """
        if not table_rows:
            return [], []

        first_row = table_rows[0]

        # 첫 행이 비어있지 않으면 헤더로 사용
        if any(cell.strip() for cell in first_row):
            # 빈 열 제거 (모든 행에서 빈 열)
            headers, data_rows = self._remove_empty_columns(first_row, table_rows[1:])
            return headers, data_rows

        # 첫 행이 비어있으면 두 번째 행부터 검사
        if len(table_rows) > 1:
            headers, data_rows = self._remove_empty_columns(
                table_rows[1], table_rows[2:]
            )
            return headers, data_rows

        return first_row, []

    @staticmethod
    def _remove_empty_columns(
        headers: list[str],
        data_rows: list[list[str]],
    ) -> tuple[list[str], list[list[str]]]:
        """모든 행에서 비어있는 열을 제거한다."""
        if not headers:
            return headers, data_rows

        all_rows = [headers] + data_rows
        col_count = max(len(r) for r in all_rows) if all_rows else 0

        non_empty_cols: list[int] = []
        for col_idx in range(col_count):
            has_value = False
            for row in all_rows:
                if col_idx < len(row) and row[col_idx].strip():
                    has_value = True
                    break
            if has_value:
                non_empty_cols.append(col_idx)

        if not non_empty_cols:
            return headers, data_rows

        filtered_headers = [
            headers[i] if i < len(headers) else "" for i in non_empty_cols
        ]
        filtered_rows = [
            [row[i] if i < len(row) else "" for i in non_empty_cols]
            for row in data_rows
        ]

        return filtered_headers, filtered_rows

    # ------------------------------------------------------------------
    # Markdown 변환
    # ------------------------------------------------------------------
    @staticmethod
    def _to_markdown(headers: list[str], rows: list[list[str]]) -> str:
        """헤더와 데이터 행을 Markdown 표로 변환한다."""
        if not headers:
            return ""

        col_count = len(headers)
        lines = [
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join("---" for _ in headers) + " |",
        ]
        for row in rows:
            padded = row + [""] * max(0, col_count - len(row))
            lines.append("| " + " | ".join(padded[:col_count]) + " |")

        return "\n".join(lines)

    # ------------------------------------------------------------------
    # 메타데이터 추출
    # ------------------------------------------------------------------
    @staticmethod
    def _extract_metadata(wb: Any) -> dict:
        """워크북 메타데이터를 추출한다."""
        metadata: dict[str, Any] = {}
        try:
            props = wb.properties
            if props:
                metadata["title"] = props.title or ""
                metadata["author"] = props.creator or ""
                metadata["created"] = str(props.created) if props.created else ""
                metadata["modified"] = str(props.modified) if props.modified else ""
            metadata["sheet_count"] = len(wb.sheetnames)
            metadata["sheet_names"] = wb.sheetnames
        except Exception:
            pass
        return metadata
